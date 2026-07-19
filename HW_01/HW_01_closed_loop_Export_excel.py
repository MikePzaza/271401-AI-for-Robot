"""
HW_01: จำลองระบบควบคุมอุณหภูมิห้องด้วยแอร์แบบ Closed Loop
Transfer Function: G(s) = K / (tau*s + 1)
Closed Loop Transfer Function: T(s) = G(s) / (1 + G(s)*H(s)), H(s) = 1

คำถาม:
1. ถ้าเปลี่ยน K เป็น 1 ถึง 10 จะส่งผลต่อระบบอย่างไร?
2. ถ้าเพิ่ม tau เป็น 2 ถึง 3 เท่า อุณหภูมิจะเปลี่ยนแปลงช้าหรือเร็วขึ้น หรือเกิดอะไรขึ้น?
3. ทำไมระบบนี้จึงเรียกว่า Closed Loop?
อย่าลืม pip install control matplotlib numpy pandas openpyxl
"""

import os
import numpy as np
import matplotlib.pyplot as plt
import control as ctrl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import LineChart, Reference

# ---------------------------------------------------------
# กำหนดโฟลเดอร์ที่จะ save ไฟล์ผลลัพธ์ทั้งหมด
# ให้ save ไว้ที่โฟลเดอร์เดียวกับตัวสคริปต์นี้เสมอ
# ไม่ว่าจะรันจาก terminal, Run button, หรือ Code Runner ก็ตาม
# ---------------------------------------------------------
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
print(f"ไฟล์ผลลัพธ์จะถูกบันทึกไว้ที่: {OUTPUT_DIR}")

# ---------------------------------------------------------
# ฟังก์ชันสร้างระบบ Closed Loop จากค่า K และ tau ที่กำหนด
# ---------------------------------------------------------
def closed_loop_system(K, tau):
    """
    สร้าง Transfer Function แบบ Closed Loop
    G(s) = K / (tau*s + 1)   -> Plant (แอร์)
    H(s) = 1                 -> Feedback (sensor วัดอุณหภูมิจริง)
    T(s) = G(s) / (1 + G(s)*H(s))
    """
    G = ctrl.tf([K], [tau, 1])   # G(s) = K / (tau*s + 1)
    H = ctrl.tf([1], [1])        # H(s) = 1 (unity feedback)
    T = ctrl.feedback(G, H)      # T(s) = G / (1 + G*H)
    return T


# ---------------------------------------------------------
# ค่าเริ่มต้น (Base case)
# ---------------------------------------------------------
K_base = 5      # กำลังทำความเย็นของแอร์
tau_base = 1     # ความเฉื่อยของระบบ (วินาที/นาที)

t = np.linspace(0, 10, 500)

# ===========================================================
# ส่วนที่ 1: ทดลองเปลี่ยนค่า K = 1 ถึง 10 (tau คงที่)
# ===========================================================
plt.figure(figsize=(8, 5))
K_values = [1, 2, 4, 6, 8, 10]
df_K = pd.DataFrame({"Time_s": t})

for K in K_values:
    T = closed_loop_system(K, tau_base)
    time, response = ctrl.step_response(T, t)
    plt.plot(time, response, label=f"K = {K}")
    df_K[f"K={K}"] = response

plt.title(f"Step Response - Varying K (tau = {tau_base})")
plt.xlabel("Time (s)")
plt.ylabel("Room Temperature Response")
plt.axhline(1, color="gray", linestyle="--", linewidth=0.8, label="Setpoint (Normalized = 1)")
plt.legend()
plt.grid(True)
plt.tight_layout()
K_png_path = os.path.join(OUTPUT_DIR, "response_vary_K.png")
plt.savefig(K_png_path, dpi=150)
print(f"บันทึกกราฟแล้ว: {K_png_path}")
plt.show()

# ===========================================================
# ส่วนที่ 2: ทดลองเปลี่ยนค่า tau = 1, 2, 3 เท่า (K คงที่)
# ===========================================================
plt.figure(figsize=(8, 5))
tau_values = [1, 2, 3]
df_tau = pd.DataFrame({"Time_s": t})

for tau in tau_values:
    T = closed_loop_system(K_base, tau)
    time, response = ctrl.step_response(T, t)
    plt.plot(time, response, label=f"tau = {tau}")
    df_tau[f"tau={tau}"] = response

plt.title(f"Step Response - Varying tau (K = {K_base})")
plt.xlabel("Time (s)")
plt.ylabel("Room Temperature Response")
plt.axhline(K_base/(1+K_base), color="gray", linestyle="--", linewidth=0.8, label="Steady State Value")
plt.legend()
plt.grid(True)
plt.tight_layout()
tau_png_path = os.path.join(OUTPUT_DIR, "response_vary_tau.png")
plt.savefig(tau_png_path, dpi=150)
print(f"บันทึกกราฟแล้ว: {tau_png_path}")
plt.show()

# ===========================================================
# ส่วนที่ 3: วิเคราะห์และพิมพ์คำตอบ
# ===========================================================
print("=" * 60)
print("คำตอบคำถามท้าย HW_01")
print("=" * 60)

print("""
1) ถ้าเปลี่ยน K เป็น 1 ถึง 10 จะส่งผลต่อระบบอย่างไร?
   - เมื่อ K มากขึ้น ระบบตอบสนองเร็วขึ้นและค่า Steady State
     เข้าใกล้ Setpoint (ค่า 1) มากขึ้น เพราะ Steady State =
     K / (1 + K) ซึ่งจะเข้าใกล้ 1 เมื่อ K มีค่าสูงมาก
   - อย่างไรก็ตาม ถ้า K สูงเกินไปในระบบจริง อาจทำให้เกิด
     Overshoot หรือแอร์ทำงานหนักเกินไป (แม้ในระบบ First-order
     นี้จะไม่มี Overshoot ก็ตาม เพราะไม่มีขั้ว complex)

2) ถ้าเพิ่ม tau เป็น 2 ถึง 3 เท่า อุณหภูมิจะเปลี่ยนแปลงช้าลง
   เพราะ tau คือค่าคงที่เวลา (Time Constant) ของระบบ
   ยิ่ง tau มาก ระบบยิ่งตอบสนองช้า (ห้องใหญ่ขึ้น หรือมี inertia
   สูงขึ้น) ทำให้ใช้เวลานานขึ้นกว่าจะเข้าสู่ Steady State
   Time constant ใหม่ของระบบวงปิดคือ tau' = tau / (1+K)

3) ทำไมระบบนี้จึงเรียกว่า Closed Loop?
   เพราะมีการวัดค่าอุณหภูมิจริงจาก sensor (Feedback, H(s))
   แล้วนำค่านั้นย้อนกลับมาเปรียบเทียบกับ Setpoint (Error
   Detection) เพื่อปรับการทำงานของแอร์อย่างต่อเนื่อง ทำให้ระบบ
   สามารถชดเชยผลกระทบจากสิ่งรบกวน (Disturbance) ได้ ต่างจาก
   Open Loop ที่ไม่มีการนำเอาต์พุตย้อนกลับมาตรวจสอบเลย
""")

# ===========================================================
# ส่วนที่ 4: Export ข้อมูลเป็นไฟล์ Excel
# ===========================================================
def write_df(ws, df, title):
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    header_row = 3
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for row_i, row in enumerate(df.itertuples(index=False), start=header_row + 1):
        for col_i, val in enumerate(row, start=1):
            c = ws.cell(row=row_i, column=col_i, value=round(val, 5))
            c.font = Font(name="Arial")
    for col_idx in range(1, len(df.columns) + 1):
        ws.column_dimensions[ws.cell(row=header_row, column=col_idx).column_letter].width = 12
    return header_row


def add_chart(ws, df, hdr_row, title, anchor):
    chart = LineChart()
    chart.title = title
    chart.x_axis.title = "Time (s)"
    chart.y_axis.title = "Response"
    n_rows = len(df)
    n_cols = len(df.columns)
    cats = Reference(ws, min_col=1, min_row=hdr_row + 1, max_row=hdr_row + n_rows)
    for col in range(2, n_cols + 1):
        data = Reference(ws, min_col=col, min_row=hdr_row, max_row=hdr_row + n_rows)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 18
    chart.height = 10
    ws.add_chart(chart, anchor)


wb = Workbook()

ws1 = wb.active
ws1.title = "Vary_K"
hdr1 = write_df(ws1, df_K, f"Step Response - Varying K (tau = {tau_base})")
add_chart(ws1, df_K, hdr1, "Step Response - Varying K", "I3")

ws2 = wb.create_sheet("Vary_tau")
hdr2 = write_df(ws2, df_tau, f"Step Response - Varying tau (K = {K_base})")
add_chart(ws2, df_tau, hdr2, "Step Response - Varying tau", "I3")

ws3 = wb.create_sheet("Analysis_Answers")
ws3["A1"] = "HW_01: Closed Loop Temperature Control - Analysis"
ws3["A1"].font = Font(name="Arial", size=14, bold=True)

answers = [
    ("1) Effect of K (1 to 10):",
     "As K increases, the system responds faster and the steady-state value "
     "approaches the setpoint (Steady State = K/(1+K)). Higher K brings the "
     "output closer to 1 (the normalized setpoint)."),
    ("2) Effect of tau (2x to 3x):",
     "As tau increases, the system responds more slowly (larger room / higher "
     "inertia). The new closed-loop time constant is tau' = tau/(1+K), so a "
     "larger tau directly slows down the temperature change."),
    ("3) Why is this called Closed Loop?",
     "Because the actual room temperature is measured by a sensor (feedback, H(s)) "
     "and compared with the setpoint (error detection) to continuously adjust the "
     "air-conditioner. This feedback lets the system compensate for disturbances, "
     "unlike an Open Loop system which has no feedback."),
]
row = 3
for q, a in answers:
    c = ws3.cell(row=row, column=1, value=q)
    c.font = Font(name="Arial", bold=True)
    ws3.cell(row=row + 1, column=1, value=a).font = Font(name="Arial")
    ws3.cell(row=row + 1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws3.row_dimensions[row + 1].height = 60
    row += 3

ws3.column_dimensions["A"].width = 100

excel_path = os.path.join(OUTPUT_DIR, "HW_01_ClosedLoop_Data.xlsx")
wb.save(excel_path)
print(f"Excel file saved: {excel_path}")
